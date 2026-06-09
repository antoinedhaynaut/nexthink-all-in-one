"""Convert platform_content.xlsx → data.json for the infographic.

Usage:
    python xlsx_to_json.py [input.xlsx] [output.json]

Defaults: platform_content.xlsx → data.json (next to it).

Validates referential integrity:
- every use_case.value_category_id exists in value_categories
- every use_case_features.{use_case_id, feature_id} exists
- every use_case_personas.{use_case_id, persona_id} exists
- every use_case_benefits.use_case_id exists

Exits non-zero on any validation error.
"""
import json
import re
import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

DEFAULT_IN   = "platform_content.xlsx"
DEFAULT_OUT  = "data.json"
DEFAULT_HTML = "platform_infographic.html"

BUNDLE_START = "/*BUNDLED_DATA_START*/"
BUNDLE_END   = "/*BUNDLED_DATA_END*/"

def read_sheet(wb, name):
    """Return a list of dict rows. Header row = first row whose first cell is not italic-styled note."""
    if name not in wb.sheetnames:
        raise SystemExit(f"missing sheet: {name}")
    ws = wb[name]
    header_row_idx = None
    for r in range(1, ws.max_row + 1):
        first = ws.cell(row=r, column=1)
        # notes rows are italic; header row is bold non-italic
        if first.font and first.font.bold and not first.font.italic:
            header_row_idx = r
            break
    if header_row_idx is None:
        # fallback: row 1
        header_row_idx = 1
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_idx, column=c).value
        if v is None:
            break
        headers.append(str(v).strip())
    rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row = {}
        empty = True
        for ci, h in enumerate(headers, 1):
            v = ws.cell(row=r, column=ci).value
            if v is not None and v != "":
                empty = False
            row[h] = v
        if not empty:
            rows.append(row)
    return rows

def coerce_int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except Exception:
            return default

def main():
    in_path  = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_IN)
    out_path = Path(sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT)
    if not in_path.exists():
        raise SystemExit(f"input not found: {in_path}")
    wb = load_workbook(in_path, data_only=True)

    sources    = read_sheet(wb, "data_sources")
    features   = read_sheet(wb, "foundation_features")
    vcats      = read_sheet(wb, "value_categories")
    ucases     = read_sheet(wb, "use_cases")
    uc_feat    = read_sheet(wb, "use_case_features")
    personas   = read_sheet(wb, "personas")
    uc_pers    = read_sheet(wb, "use_case_personas")
    uc_benefits= read_sheet(wb, "use_case_benefits")
    uc_stories = read_sheet(wb, "use_case_stories") if "use_case_stories" in wb.sheetnames else []

    # Build lookup sets for validation
    feat_ids   = {r["id"] for r in features}
    vcat_ids   = {r["id"] for r in vcats}
    uc_ids     = {r["id"] for r in ucases}
    pers_ids   = {r["id"] for r in personas}

    errors = []

    # validate use_cases ↔ value_categories
    for u in ucases:
        if u.get("value_category_id") not in vcat_ids:
            errors.append(f"use_case '{u.get('id')}' has unknown value_category_id '{u.get('value_category_id')}'")

    # build M:N maps
    uc_to_features  = defaultdict(list)
    for r in uc_feat:
        if r["use_case_id"] not in uc_ids:
            errors.append(f"use_case_features references unknown use_case '{r['use_case_id']}'")
        if r["feature_id"] not in feat_ids:
            errors.append(f"use_case_features references unknown feature '{r['feature_id']}'")
        uc_to_features[r["use_case_id"]].append(r["feature_id"])

    uc_to_personas = defaultdict(list)
    for r in uc_pers:
        if r["use_case_id"] not in uc_ids:
            errors.append(f"use_case_personas references unknown use_case '{r['use_case_id']}'")
        if r["persona_id"] not in pers_ids:
            errors.append(f"use_case_personas references unknown persona '{r['persona_id']}'")
        uc_to_personas[r["use_case_id"]].append(r["persona_id"])

    uc_to_stories = defaultdict(list)
    for r in uc_stories:
        if r.get("use_case_id") not in uc_ids:
            errors.append(f"use_case_stories references unknown use_case '{r.get('use_case_id')}'")
        else:
            uc_to_stories[r["use_case_id"]].append({
                "title":    r.get("title") or "",
                "url":      r.get("url") or "",
                "customer": r.get("customer") or "Anonymous",
                "industry": r.get("industry") or "",
                "order":    coerce_int(r.get("order"), 999),
            })

    uc_to_benefits = defaultdict(list)
    for r in uc_benefits:
        if r.get("use_case_id") not in uc_ids:
            errors.append(f"use_case_benefits references unknown use_case '{r.get('use_case_id')}'")
        else:
            uc_to_benefits[r["use_case_id"]].append(r.get("benefit_type") or "")

    if errors:
        for e in errors: print("ERROR:", e, file=sys.stderr)
        raise SystemExit(f"{len(errors)} referential integrity error(s) — aborting.")

    # Sort stories by order
    for k, lst in uc_to_stories.items():
        lst.sort(key=lambda x: x["order"])

    # Compose final JSON
    out = {
        "dataSources": [
            {
                "id":          r["id"],
                "name":        r["name"],
                "category":    r.get("category") or "",
                "icon":        r.get("icon") or "",
                "description": r.get("description") or "",
                "order":       coerce_int(r.get("order"), 999),
            } for r in sorted(sources, key=lambda x: coerce_int(x.get("order"), 999))
        ],
        "foundationFeatures": [
            {
                "id":          r["id"],
                "name":        r["name"],
                "group":       (r.get("group") or "").lower(),
                "description": r.get("description") or "",
                "order":       coerce_int(r.get("order"), 999),
            } for r in sorted(features, key=lambda x: coerce_int(x.get("order"), 999))
        ],
        "valueCategories": [
            {
                "id":      r["id"],
                "name":    r["name"],
                "color":   (r.get("color") or "9CA3AF").lstrip("#"),
                "tagline": r.get("tagline") or "",
                "order":   coerce_int(r.get("order"), 999),
            } for r in sorted(vcats, key=lambda x: coerce_int(x.get("order"), 999))
        ],
        "personas": [
            {
                "id":    r["id"],
                "name":  r["name"],
                "role":  r.get("role") or "",
                "icon":  r.get("icon") or "",
                "order": coerce_int(r.get("order"), 999),
            } for r in sorted(personas, key=lambda x: coerce_int(x.get("order"), 999))
        ],
        "useCases": [
            {
                "id":              u["id"],
                "name":            u["name"],
                "valueCategoryId": u["value_category_id"],
                "shortPitch":        u.get("short_pitch") or "",
                "valueDescription": u.get("value_description") or "",
                "benefitTypes":     uc_to_benefits.get(u["id"], []),
                "customerStories":  uc_to_stories.get(u["id"], []),
                "featureIds":  uc_to_features.get(u["id"], []),
                "personaIds":  uc_to_personas.get(u["id"], []),
                "order":       coerce_int(u.get("order"), 999),
            } for u in sorted(ucases, key=lambda x: (x.get("value_category_id") or "", coerce_int(x.get("order"), 999)))
        ],
    }

    json_text = json.dumps(out, indent=2, ensure_ascii=False)
    out_path.write_text(json_text, encoding="utf-8")
    n_uc = len(out["useCases"])
    n_vc = len(out["valueCategories"])
    n_f  = len(out["foundationFeatures"])
    n_s  = len(out["dataSources"])
    n_p  = len(out["personas"])
    n_b  = sum(len(u["benefitTypes"]) for u in out["useCases"])
    n_st = sum(len(u["customerStories"]) for u in out["useCases"])
    print(f"OK → {out_path}  ({n_s} sources · {n_f} features · {n_vc} value cats · {n_uc} use cases · {n_p} personas · {n_b} benefit links · {n_st} story links)")

    # Inject the JSON inline into the HTML so the file is openable from file:// (no server)
    html_path = out_path.parent / DEFAULT_HTML
    if html_path.exists():
        html = html_path.read_text(encoding="utf-8")
        pattern = re.compile(re.escape(BUNDLE_START) + r"[\s\S]*?" + re.escape(BUNDLE_END))
        # Escape </script> to keep the embedded JSON safe inside <script>
        safe_json = json_text.replace("</", "<\\/")
        replacement = f"{BUNDLE_START}\n{safe_json}\n{BUNDLE_END}"
        if pattern.search(html):
            html = pattern.sub(replacement, html, count=1)
            html_path.write_text(html, encoding="utf-8")
            print(f"OK → {html_path} (data bundled inline)")
        else:
            print(f"WARN: could not find {BUNDLE_START}…{BUNDLE_END} markers in {html_path}; skipped inline bundle.")
    else:
        print(f"NOTE: {html_path} not found — skipping inline bundle.")

if __name__ == "__main__":
    main()

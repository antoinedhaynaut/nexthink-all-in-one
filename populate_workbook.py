"""
Rebuild platform_content.xlsx with the new schema:
 - Rewritten value_categories (same IDs, updated names/colors/taglines)
 - New clean use case IDs replacing formula-based rows
 - value_description field per use case (plain-text value narrative)
 - New sheet: use_case_benefits (use_case_id, benefit_type)
 - Cleared maturity_levels
 - Rewired use_case_features and use_case_personas

Run from the Nexthink Infographic folder:
    python3 populate_workbook.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from pathlib import Path
from copy import copy

WB_PATH = Path("platform_content.xlsx")

# ── Helpers ────────────────────────────────────────────────────────────────────

def note_style():
    return Font(italic=True, color="9CA3AF")

def header_style():
    return Font(bold=True, color="FFFFFF")

def header_fill():
    return PatternFill("solid", fgColor="1E293B")

def write_sheet(ws, headers, rows, note=""):
    """Clear sheet and write note row + header row + data rows."""
    # Remove any merged-cell ranges that survive delete_rows
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))
    ws.delete_rows(1, ws.max_row + 1)
    r = 1
    if note:
        ws.cell(r, 1, note).font = note_style()
        r += 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(r, ci, h)
        c.font = header_style()
        c.fill = header_fill()
    for row in rows:
        r += 1
        for ci, v in enumerate(row, 1):
            ws.cell(r, ci, v)

# ── Data ───────────────────────────────────────────────────────────────────────

VALUE_CATEGORIES = [
    # id, name, color (hex no #), tagline, order
    ("vc_service_mgmt", "Service Management",      "5B7FFF", "Reduce reactive workload and prevent disruptions",      10),
    ("vc_security",     "Compliance",               "0EA5E9", "Maintain posture, enforce standards, manage risk",      20),
    ("vc_workplace",    "Asset Management",         "F97316", "Right-size, optimize and rationalize IT assets",        30),
    ("vc_sustain",      "Sustainability",           "34D399", "Reduce carbon, energy and hardware procurement",        40),
    ("vc_dex",          "Strategy",                 "A855F7", "Improve experience, revenue and talent retention",      50),
    ("vc_change",       "Change & Transformation",  "3B82F6", "De-risk migrations, adoptions and M&A activity",        60),
]

USE_CASES = [
    # id, name, vc_id, short_pitch, customer, industry, oneliner, result, value_description, order
    # ── Service Management ──
    ("uc_sm_reactive",
     "Reduce reactive workload",
     "vc_service_mgmt",
     "Shift from firefighting to prevention with AI-powered issue detection and auto-remediation.",
     "Initech", "Retail",
     "Caught a faulty driver rollout 3 hours after deployment.",
     "Zero tickets opened · rollback in 25 min.",
     "Nexthink gives IT full visibility to see, diagnose and fix issues before employees even notice them. By detecting recurring problems across the fleet — from crashes and connectivity drops to application freezes — IT can shift from reactive firefighting to proactive resolution, reducing inbound ticket volumes and shortening mean time to resolution.",
     10),

    ("uc_sm_automation",
     "Proactive Automation",
     "vc_service_mgmt",
     "Automate repetitive IT tasks and runbooks to free agents for high-value work.",
     "Umbrella", "Healthcare",
     "Replaced 14 standard runbooks with auto-remediation.",
     "L1 ticket volume −38% in 4 months.",
     "Nexthink automates the detection and resolution of recurring issues and problems at scale, utilizing self-help and self-heal capabilities. Shifting to proactive IT improves operational efficiency and reduces the cost of delivering IT services, freeing up service desk resources for higher-value work.",
     20),

    ("uc_sm_efficiency",
     "Improve IT efficiency",
     "vc_service_mgmt",
     "Measure and benchmark IT operations to eliminate waste and accelerate resolution.",
     "GlobalTech", "Manufacturing",
     "Consolidated 3 IT dashboards into a single pane of glass.",
     "MTTR reduced by 42% in 6 months.",
     "Nexthink data integrated with ITSM tooling reduces the time employees spend on the phone with IT, while enhanced visibility enables IT to fix non-reported issues across the enterprise. Enhancing ITIL service and business processes with DEX data drives measurable operational efficiency gains.",
     30),

    ("uc_sm_prevention",
     "Prevent technology disruptions",
     "vc_service_mgmt",
     "Detect performance regressions and hardware failures before they cause downtime.",
     "EverGreen Bank", "Financial services",
     "Prevented a storage failure from impacting 4,000 traders.",
     "Zero business downtime · $800k loss avoided.",
     "Proactive management of SaaS and collaboration tools, combined with real-time fleet monitoring, enables IT to prevent technology disruptions before they impact employees. Nexthink continuously monitors device health, application performance and network connectivity to identify degradation patterns and trigger automated fixes.",
     40),

    ("uc_sm_onboarding",
     "Employee Onboarding",
     "vc_service_mgmt",
     "Automate device setup and guided onboarding so new hires are productive on day one.",
     "Acme Corp", "Manufacturing",
     "Automated onboarding for 3,200 new hires across 12 sites.",
     "Time-to-productivity cut from 5 days to 1 day.",
     "Nexthink improves the employee onboarding process from device provisioning to application enablement. By monitoring the digital experience of new joiners, IT can ensure devices are correctly configured, applications are accessible, and onboarding campaigns guide employees through key tools and processes at the right moment.",
     50),

    ("uc_spark",
     "Self-help AI agent",
     "vc_service_mgmt",
     "Give every employee an autonomous AI agent that resolves IT issues without a ticket.",
     None, None, None, None,
     "Nexthink Spark, an agentic AI, integrates across the tech stack to resolve issues autonomously, eliminating human-touch ticket resolution. Employees describe their problem in natural language and Spark diagnoses the issue using live device telemetry, then executes IT-approved remediation actions — reducing the need to call the service desk and accelerating resolution without any agent involvement.",
     60),

    # ── Compliance ──
    ("uc_comp_sec_posture",
     "Improve security posture",
     "vc_security",
     "Continuously assess and harden endpoint security across the entire fleet.",
     "Cyberdyne", "Defense",
     "Hardened 22,000 endpoints to ISO 27001 standard.",
     "Audit prep: 6 weeks → 4 days.",
     "Nexthink provides continuous, real-time visibility into the security posture of every managed endpoint. By surfacing configuration gaps, missing patches, and policy drift across the fleet, IT security teams can proactively identify and close vulnerabilities before they are exploited — dramatically reducing audit preparation time and improving overall compliance scores.",
     10),

    ("uc_comp_standards",
     "Enforce corporate standards",
     "vc_security",
     "Detect and auto-remediate policy drift the moment it occurs.",
     "Tyrell Corp", "Tech",
     "Deployed 40 compliance policies to 30,000 endpoints.",
     "Policy compliance: 74% → 98% in 60 days.",
     "Nexthink continuously monitors endpoint configurations against corporate SOE standards and automatically remediates deviations at scale. When a device drifts from policy — through a user action, software update, or OS change — Nexthink detects it instantly and can trigger a self-healing action or alert the relevant team, maintaining compliance without manual intervention.",
     20),

    ("uc_comp_human_risk",
     "Manage Human Risk",
     "vc_security",
     "Identify risky user behaviours and intervene with targeted in-app guidance.",
     "Stark Industries", "Aerospace",
     "Identified 1,200 users bypassing security controls.",
     "Risk incidents down 55% after targeted campaigns.",
     "Nexthink monitors employee behaviours that introduce security risk — from shadow IT tool adoption to unsafe browsing and policy bypasses. When risky patterns are detected, IT can deploy targeted in-app campaigns to educate and redirect employees in the moment, reducing human-caused security incidents without heavy-handed enforcement.",
     30),

    ("uc_comp_data",
     "Data Integrity & Compliance",
     "vc_security",
     "Ensure data governance and integrity across all managed devices and systems.",
     "Wayne Enterprises", "Conglomerate",
     "Tracked data flows for GDPR compliance across 8 regions.",
     "Regulatory fines risk reduced · audit passed first time.",
     "Nexthink tracks data flows, access patterns and application usage across managed devices to support GDPR, HIPAA and other data governance frameworks. By correlating endpoint telemetry with data classification policies, IT can identify where sensitive data is being handled outside approved processes and trigger remediation before a breach occurs.",
     40),

    ("uc_comp_ai",
     "AI Compliance",
     "vc_security",
     "Govern shadow AI tool usage and enforce AI policies across the workforce.",
     "Pied Piper", "Tech",
     "Detected 47 unauthorized AI tools used by 1,800 employees.",
     "AI governance policy deployed fleet-wide in 2 weeks.",
     "As AI tool adoption accelerates, Nexthink provides full visibility into which AI applications employees are using — including unsanctioned tools. IT and security teams can monitor AI usage patterns, enforce governance policies, and run adoption campaigns for approved AI tools, ensuring responsible AI use across the workforce.",
     50),

    # ── Asset Management ──
    ("uc_am_hw_refresh",
     "Performance based HW Refresh",
     "vc_workplace",
     "Replace hardware based on real performance data, not arbitrary age policies.",
     "Globex Bank", "Financial services",
     "Identified 1,200 devices with measurable performance issues.",
     "Replaced 1,200 devices · $2.1M in support costs avoided.",
     "Nexthink replaces arbitrary age-based refresh policies with data-driven decisions. By combining real-time performance metrics — crash rates, application load times, CPU and memory utilisation — with hardware attributes and user productivity impact, IT can precisely identify devices that genuinely need replacement, avoiding both premature refreshes and the productivity drain of underperforming machines.",
     10),

    ("uc_am_hw_sizing",
     "Hardware Right-sizing",
     "vc_workplace",
     "Match hardware spec to actual workload to avoid over-provisioning and waste.",
     "Stark Industries", "Aerospace",
     "Right-sized 8,000 laptops based on actual resource usage.",
     "$3.2M CapEx avoided in one refresh cycle.",
     "Nexthink analyses actual resource consumption against device specifications to identify over- and under-provisioned hardware. By building persona-based hardware profiles grounded in real workload data, IT can right-size every device category in the fleet — ensuring new purchases match actual needs and avoiding costly over-specification.",
     20),

    ("uc_am_sw_spend",
     "Reduce Software Spend",
     "vc_workplace",
     "Eliminate unused licenses and consolidate redundant SaaS tools.",
     "Wayne Enterprises", "Conglomerate",
     "Identified 4,500 dormant E5 licenses in week 2.",
     "$1.2M annual SaaS spend reclaimed.",
     "Nexthink provides complete visibility into SaaS application usage across the enterprise, tracking which licenses are actively used, underused, or dormant. By surfacing exact utilisation data, IT and finance teams can confidently eliminate waste, consolidate overlapping tools, and negotiate renewals with leverage — turning software spend into a managed cost.",
     30),

    ("uc_am_ai_optim",
     "AI Consumption & Optimisation",
     "vc_workplace",
     "Track AI tool adoption and optimize Copilot/ChatGPT license usage.",
     "MegaCorp", "Pharma",
     "Identified 2,100 unused Copilot licenses.",
     "$630k/year in unused AI licenses reclaimed.",
     "As organisations invest heavily in AI productivity tools like Microsoft Copilot, Nexthink tracks who is actually using them, how frequently, and in which workflows. This real-time adoption data enables IT to reclaim unused licenses, redeploy them to higher-value users, and justify — or challenge — further AI investments with hard evidence.",
     40),

    ("uc_am_vdi",
     "VDI Optimisation",
     "vc_workplace",
     "Right-size VDI sessions and identify underutilized virtual machines.",
     "Initech", "Retail",
     "Reduced VDI session over-provisioning by 34%.",
     "$1.8M in VDI infrastructure costs avoided.",
     "Nexthink's VDI Experience module provides deep visibility into virtual desktop performance — session responsiveness, logon times, application latency and resource contention. By identifying over-provisioned sessions and underutilised infrastructure, IT can right-size the VDI estate and consolidate infrastructure without degrading the employee experience.",
     50),

    ("uc_am_mobile",
     "Mobile & POS",
     "vc_workplace",
     "Ensure mobile devices and point-of-sale terminals stay healthy and compliant.",
     "Greenfield Co", "Energy",
     "Monitored 6,000 mobile devices across 80 field sites.",
     "Mobile downtime reduced by 61% in 3 months.",
     "Nexthink extends real-time health monitoring to mobile devices and point-of-sale terminals, tracking battery health, connectivity quality, application performance and compliance status. IT teams gain the same level of visibility into mobile fleets as traditional endpoints, enabling proactive intervention before device issues impact frontline workers or customer-facing operations.",
     60),

    # ── Sustainability ──
    ("uc_sus_carbon",
     "Reduce Carbon Footprint",
     "vc_sustain",
     "Measure, report and reduce the carbon impact of the device fleet.",
     "Greenfield Co", "Energy",
     "Cut idle laptop power by 38% via policy and nudges.",
     "−540 t CO₂eq/yr · ESG data automated.",
     "Nexthink identifies idle and always-on devices that consume energy unnecessarily — typically 25–30% of the fleet. By automating shutdown campaigns and smart power policies, IT can reduce the fleet's energy consumption and calculate the resulting carbon savings with precision, providing ESG-ready data for sustainability reporting.",
     10),

    ("uc_sus_energy",
     "Reduce Energy costs",
     "vc_sustain",
     "Optimize sleep policies and identify always-on devices to cut energy spend.",
     "EverGreen Bank", "Financial services",
     "Deployed smart power policies to 22,000 endpoints.",
     "1.2 GWh/yr saved · €180k annual energy saving.",
     "Nexthink monitors device power states across the fleet and automatically enforces energy-efficient sleep and shutdown policies. By targeting specific device populations — those left on overnight, those with aggressive display settings — IT can reduce annual energy consumption significantly and provide Finance with auditable cost savings.",
     20),

    ("uc_sus_hw_proc",
     "Reduced hardware procurement",
     "vc_sustain",
     "Extend device lifecycles using real performance data to defer CapEx.",
     "Acme Corp", "Manufacturing",
     "Extended average laptop lifecycle from 3.1 to 4.4 years.",
     "$8.4M CapEx deferred over 2 years.",
     "With 80% of a device's lifecycle carbon emissions occurring during manufacturing, extending hardware lifespan is the single highest-impact sustainability action IT can take. Nexthink uses real performance and health data to identify devices that can safely continue in service, enabling IT to defer procurement and provide tangible sustainability metrics to stakeholders.",
     30),

    # ── Strategy ──
    ("uc_str_experience",
     "Improve employee & customer Experience",
     "vc_dex",
     "Measure and improve the digital experience of every employee with a unified DEX score.",
     "Acme Corp", "Manufacturing",
     "Rolled out DEX scoring across 18 countries in 6 weeks.",
     "+12 DEX pts in 5 months · 30% fewer P1 tickets.",
     "Nexthink's DEX Score aggregates thousands of endpoint signals — performance, application reliability, connectivity quality and employee sentiment — into a single actionable measure of digital experience. IT leaders can track experience trends across the enterprise, segment by persona, location or device type, and correlate DEX improvements with business outcomes.",
     10),

    ("uc_str_revenue",
     "Increase Revenue",
     "vc_dex",
     "Link employee digital experience to business productivity and revenue outcomes.",
     "Globex Bank", "Financial services",
     "Correlated DEX score with trader performance metrics.",
     "Top-quartile DEX users generated 18% more revenue.",
     "Research consistently shows that employees with better digital experiences are more productive and more engaged. Nexthink enables organisations to quantify this link — correlating DEX scores with productivity metrics, output quality and revenue per employee — building the business case for IT investment in terms that resonate with the CFO and CEO.",
     20),

    ("uc_str_data",
     "Become Data Driven",
     "vc_dex",
     "Build a data platform for IT decisions that eliminates guesswork and opinion.",
     "Pied Piper", "Tech",
     "Replaced manual reporting with live IT intelligence dashboards.",
     "Decision cycle time cut from 2 weeks to 2 hours.",
     "Nexthink transforms IT from a cost centre into a data-driven business partner by providing contextual, real-time data on the digital workplace. By integrating endpoint intelligence with ITSM, HR and business systems, IT leaders can make strategic decisions about technology investments, workforce productivity and operational risk based on evidence rather than opinion.",
     30),

    ("uc_str_talent",
     "Attract & Retain Talent",
     "vc_dex",
     "Use digital experience as a talent differentiator to reduce attrition.",
     "Wayne Enterprises", "Conglomerate",
     "Launched IT satisfaction programme tied to employer brand.",
     "IT-related attrition down 22% · eNPS +14.",
     "Digital experience is increasingly a talent differentiator — employees cite poor technology as a top reason for leaving. Nexthink enables IT to measure, track and improve the employee technology experience at scale, giving HR and leadership teams a credible story for employer branding, onboarding quality and reducing IT-related attrition.",
     40),

    ("uc_str_xla",
     "XLA Methodology",
     "vc_dex",
     "Shift from SLA compliance to experience-level agreements that measure what users actually feel.",
     "MegaCorp", "Pharma",
     "Replaced 6 SLA dashboards with 1 XLA scorecard.",
     "Executive NPS +18 · XLA adoption in 3 business units.",
     "Experience Level Agreements shift IT's accountability from measuring response times to measuring the quality of experience employees actually have. Nexthink provides the continuous observability layer that makes XLAs operational — combining performance telemetry, sentiment data and proactive issue detection to give IT teams a real-time view of whether they are delivering on experience commitments.",
     50),

    # ── Change & Transformation ──
    ("uc_ct_budget",
     "Budgeting and planning",
     "vc_change",
     "Use real device and usage data to plan IT budgets with confidence.",
     "Stark Industries", "Aerospace",
     "Built a data-driven 3-year IT investment model.",
     "Budget variance reduced from 18% to 4%.",
     "Nexthink provides IT leaders with the endpoint data they need to build credible, evidence-based IT budgets. Hardware age, performance degradation curves, application usage trends and support cost patterns all feed into more accurate refresh planning, license forecasting and infrastructure investment decisions — reducing variance and improving stakeholder confidence in IT spending.",
     10),

    ("uc_ct_transform",
     "Accelerate Transformation",
     "vc_change",
     "Track rollout velocity, adoption and experience impact for every transformation programme.",
     "Initech", "Retail",
     "Ran a 90-day digital transformation programme across 12,000 users.",
     "Adoption target hit 3 weeks early · zero P1 incidents.",
     "When rolling out new technology — whether a cloud migration, desktop OS upgrade or collaboration platform switch — Nexthink provides real-time visibility into adoption velocity, experience impact and support demand. IT can run controlled pilots, measure the actual experience delta, and course-correct before full deployment, de-risking transformation programmes.",
     20),

    ("uc_ct_ai_adopt",
     "AI & Digital Adoption",
     "vc_change",
     "Measure and boost adoption of AI tools and new digital capabilities.",
     "Cyberdyne", "Defense",
     "Tracked Copilot adoption across 8,000 knowledge workers.",
     "Adoption from 31% to 78% in 8 weeks.",
     "Deploying AI tools is only valuable if employees use them effectively. Nexthink tracks AI adoption at the individual, team and organisational level — identifying adoption barriers, running targeted enablement campaigns for low-usage cohorts, and measuring the productivity impact of AI tools to maximise the return on AI investments.",
     30),

    ("uc_ct_mna",
     "M&A Activity",
     "vc_change",
     "Accelerate IT integration after mergers and acquisitions with full estate visibility.",
     "Wayne Enterprises", "Conglomerate",
     "Mapped and standardized IT estate of an acquired company in 6 weeks.",
     "Integration timeline cut by 40% · no business disruption.",
     "During mergers and acquisitions, IT integration is a critical path item. Nexthink provides immediate visibility into the acquired organisation's technology estate — hardware inventory, software usage, performance baselines and compliance posture — enabling IT to plan integration, identify rationalisation opportunities, and execute the transition without disrupting business operations.",
     40),
]

# ── IT & Business benefits (M:N) ────────────────────────────────────────────────
ALL_BENEFITS = [
    "IT cost avoided", "IT time saved", "IT costs saved",
    "Business time saved", "Business costs saved", "Risk avoidance",
    "Company reputation", "Overall employee experience", "Business productivity",
]

UC_BENEFITS = [
    # Service Management
    ("uc_sm_reactive",    "IT time saved"),
    ("uc_sm_reactive",    "IT costs saved"),
    ("uc_sm_reactive",    "Business productivity"),
    ("uc_sm_automation",  "IT time saved"),
    ("uc_sm_automation",  "IT costs saved"),
    ("uc_sm_automation",  "Business productivity"),
    ("uc_sm_efficiency",  "IT time saved"),
    ("uc_sm_efficiency",  "IT costs saved"),
    ("uc_sm_prevention",  "IT costs saved"),
    ("uc_sm_prevention",  "Business productivity"),
    ("uc_sm_prevention",  "Risk avoidance"),
    ("uc_sm_onboarding",  "IT time saved"),
    ("uc_sm_onboarding",  "Overall employee experience"),
    ("uc_sm_onboarding",  "Business productivity"),
    ("uc_spark",          "IT costs saved"),
    ("uc_spark",          "IT time saved"),
    ("uc_spark",          "Business time saved"),
    ("uc_spark",          "Business productivity"),
    ("uc_spark",          "Overall employee experience"),
    # Compliance
    ("uc_comp_sec_posture","Risk avoidance"),
    ("uc_comp_sec_posture","IT costs saved"),
    ("uc_comp_sec_posture","Company reputation"),
    ("uc_comp_standards",  "Risk avoidance"),
    ("uc_comp_standards",  "IT costs saved"),
    ("uc_comp_human_risk", "Risk avoidance"),
    ("uc_comp_human_risk", "Company reputation"),
    ("uc_comp_data",       "Risk avoidance"),
    ("uc_comp_data",       "IT costs saved"),
    ("uc_comp_data",       "Company reputation"),
    ("uc_comp_ai",         "Risk avoidance"),
    ("uc_comp_ai",         "Company reputation"),
    # Asset Management
    ("uc_am_hw_refresh",   "IT cost avoided"),
    ("uc_am_hw_refresh",   "Overall employee experience"),
    ("uc_am_hw_sizing",    "IT cost avoided"),
    ("uc_am_hw_sizing",    "IT costs saved"),
    ("uc_am_sw_spend",     "IT cost avoided"),
    ("uc_am_sw_spend",     "IT costs saved"),
    ("uc_am_ai_optim",     "IT cost avoided"),
    ("uc_am_ai_optim",     "IT costs saved"),
    ("uc_am_vdi",          "IT cost avoided"),
    ("uc_am_vdi",          "IT costs saved"),
    ("uc_am_vdi",          "Overall employee experience"),
    ("uc_am_mobile",       "IT cost avoided"),
    ("uc_am_mobile",       "Overall employee experience"),
    # Sustainability
    ("uc_sus_carbon",      "Business costs saved"),
    ("uc_sus_carbon",      "Company reputation"),
    ("uc_sus_energy",      "IT costs saved"),
    ("uc_sus_energy",      "Business costs saved"),
    ("uc_sus_hw_proc",     "IT cost avoided"),
    ("uc_sus_hw_proc",     "Business costs saved"),
    ("uc_sus_hw_proc",     "Company reputation"),
    # Strategy
    ("uc_str_experience",  "Overall employee experience"),
    ("uc_str_experience",  "Business productivity"),
    ("uc_str_experience",  "Company reputation"),
    ("uc_str_revenue",     "Business productivity"),
    ("uc_str_revenue",     "Business costs saved"),
    ("uc_str_data",        "Business productivity"),
    ("uc_str_data",        "IT time saved"),
    ("uc_str_talent",      "Overall employee experience"),
    ("uc_str_talent",      "Company reputation"),
    ("uc_str_xla",         "Overall employee experience"),
    ("uc_str_xla",         "Business productivity"),
    # Change & Transformation
    ("uc_ct_budget",       "IT time saved"),
    ("uc_ct_budget",       "Business costs saved"),
    ("uc_ct_transform",    "IT time saved"),
    ("uc_ct_transform",    "Business productivity"),
    ("uc_ct_ai_adopt",     "Business productivity"),
    ("uc_ct_ai_adopt",     "IT time saved"),
    ("uc_ct_mna",          "IT time saved"),
    ("uc_ct_mna",          "Business productivity"),
    ("uc_ct_mna",          "IT costs saved"),
]

# ── Customer stories (M:N, up to 3 per UC) ────────────────────────────────────
# Sources: nexthink.com/resource, nexthink.com/blog (all publicly available)
# "Anonymous" = customer name was not disclosed in the original story
UC_STORIES = [
    # (uc_id, title, url, customer, industry, order)

    # ── Service Management ──
    ("uc_sm_reactive",   "Western Union: From Reactive to Proactive",                          "https://nexthink.com/en/resource/western-union-testimonial",                                                           "Western Union",     "Financial Services",    1),
    ("uc_sm_reactive",   "Fortune 500 Healthcare Cracks Their IT Problems Once and For All",   "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            2),
    ("uc_sm_reactive",   "How Leidos Is Building a Proactive, AI-Enabled IT Organization",     "https://nexthink.com/resource/how-leidos-is-building-a-proactive-ai-enabled-it-organization",                         "Leidos",            "Defense & Technology",  3),

    ("uc_sm_automation", "Leidos Builds a Proactive IT Strategy with Nexthink",                "https://nexthink.com/resource/leidos-builds-a-proactive-it-strategy-with-nexthink",                                   "Leidos",            "Defense & Technology",  1),
    ("uc_sm_automation", "How One Company Saved with a Single Self-Help Automation",           "https://nexthink.com/blog/it-self-help-reduce-tickets/",                                                               "Anonymous",         "Manufacturing",         2),
    ("uc_sm_automation", "5 Times Nexthink Helped Customers Reduce IT Costs",                  "https://nexthink.com/blog/5-times-nexthink-helped-customers-reduce-it-costs",                                          "Multiple Customers","Various",               3),

    ("uc_sm_efficiency", "A Year in the Life of a Nexthink Customer",                         "https://nexthink.com/en/resource/a-year-in-the-life-of-a-nexthink-customer",                                           "Anonymous",         "Financial Services",    1),
    ("uc_sm_efficiency", "Johnson & Johnson: A Frictionless IT Experience",                    "https://nexthink.com/resource/johnson-johnson-a-frictionless-it-experience",                                           "Johnson & Johnson", "Healthcare",            2),
    ("uc_sm_efficiency", "Fortune 500 Healthcare Cracks Their IT Problems Once and For All",   "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            3),

    ("uc_sm_prevention", "Western Union: From Reactive to Proactive",                          "https://nexthink.com/en/resource/western-union-testimonial",                                                           "Western Union",     "Financial Services",    1),
    ("uc_sm_prevention", "How Leidos Is Building a Proactive, AI-Enabled IT Organization",     "https://nexthink.com/resource/how-leidos-is-building-a-proactive-ai-enabled-it-organization",                         "Leidos",            "Defense & Technology",  2),
    ("uc_sm_prevention", "Fortune 500 Healthcare Cracks Their IT Problems Once and For All",   "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            3),

    ("uc_sm_onboarding", "Fortune 500 Healthcare Cracks Their IT Problems Once and For All",   "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            1),
    ("uc_sm_onboarding", "Johnson & Johnson: A Frictionless IT Experience",                    "https://nexthink.com/resource/johnson-johnson-a-frictionless-it-experience",                                           "Johnson & Johnson", "Healthcare",            2),

    ("uc_spark",         "How One Company Saved with a Single Self-Help Automation",           "https://nexthink.com/blog/it-self-help-reduce-tickets/",                                                               "Anonymous",         "Manufacturing",         1),
    ("uc_spark",         "How Leidos Is Building a Proactive, AI-Enabled IT Organization",     "https://nexthink.com/resource/how-leidos-is-building-a-proactive-ai-enabled-it-organization",                         "Leidos",            "Defense & Technology",  2),
    ("uc_spark",         "5 Times Nexthink Helped Customers Reduce IT Costs",                  "https://nexthink.com/blog/5-times-nexthink-helped-customers-reduce-it-costs",                                          "Multiple Customers","Various",               3),

    # ── Compliance ──
    ("uc_comp_sec_posture","AXA Unlocks the Power of Enterprise IT Visibility",                "https://nexthink.com/en/resource/axa-unlocks-the-power-of-enterprise-it-visibility",                                   "AXA",               "Insurance",             1),
    ("uc_comp_sec_posture","Leidos Builds a Proactive IT Strategy with Nexthink",              "https://nexthink.com/resource/leidos-builds-a-proactive-it-strategy-with-nexthink",                                   "Leidos",            "Defense & Technology",  2),
    ("uc_comp_sec_posture","Fortune 500 Healthcare Cracks Their IT Problems Once and For All", "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            3),

    ("uc_comp_standards", "AXA Unlocks the Power of Enterprise IT Visibility",                 "https://nexthink.com/en/resource/axa-unlocks-the-power-of-enterprise-it-visibility",                                   "AXA",               "Insurance",             1),
    ("uc_comp_standards", "How Leidos Is Building a Proactive, AI-Enabled IT Organization",    "https://nexthink.com/resource/how-leidos-is-building-a-proactive-ai-enabled-it-organization",                         "Leidos",            "Defense & Technology",  2),

    ("uc_comp_human_risk","AXA Unlocks the Power of Enterprise IT Visibility",                 "https://nexthink.com/en/resource/axa-unlocks-the-power-of-enterprise-it-visibility",                                   "AXA",               "Insurance",             1),
    ("uc_comp_human_risk","Leidos Builds a Proactive IT Strategy with Nexthink",               "https://nexthink.com/resource/leidos-builds-a-proactive-it-strategy-with-nexthink",                                   "Leidos",            "Defense & Technology",  2),

    ("uc_comp_data",      "AXA Unlocks the Power of Enterprise IT Visibility",                 "https://nexthink.com/en/resource/axa-unlocks-the-power-of-enterprise-it-visibility",                                   "AXA",               "Insurance",             1),
    ("uc_comp_data",      "Fortune 500 Healthcare Cracks Their IT Problems Once and For All",  "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            2),

    ("uc_comp_ai",        "Secure AI Adoption at Scale with Nexthink Adopt",                   "https://nexthink.com/resource/empower-smarter-safer-ai-adoption-across-your-enterprise",                               "Anonymous",         "Utilities",             1),
    ("uc_comp_ai",        "How Leidos Is Building a Proactive, AI-Enabled IT Organization",    "https://nexthink.com/resource/how-leidos-is-building-a-proactive-ai-enabled-it-organization",                         "Leidos",            "Defense & Technology",  2),

    # ── Asset Management ──
    ("uc_am_hw_refresh",  "Hospital Saves $900k on Hardware Refresh",                          "https://nexthink.com/blog/5-times-nexthink-helped-customers-reduce-it-costs",                                          "Anonymous",         "Healthcare",            1),
    ("uc_am_hw_refresh",  "Honeywell: Running a Cost-Efficient Workplace",                     "https://nexthink.com/resource/honeywell-cost-efficient-workplace",                                                      "Honeywell",         "Manufacturing",         2),
    ("uc_am_hw_refresh",  "Extend Device Lifecycles and Increase Employee Happiness",          "https://nexthink.com/blog/extend-hardware-refresh-cycle",                                                               "Anonymous",         "Various",               3),

    ("uc_am_hw_sizing",   "German Manufacturer Saves $1.6M on Windows Migration",              "https://nexthink.com/blog/migration-savings/",                                                                          "Anonymous",         "Manufacturing",         1),
    ("uc_am_hw_sizing",   "5 Times Nexthink Helped Customers Reduce IT Costs",                 "https://nexthink.com/blog/5-times-nexthink-helped-customers-reduce-it-costs",                                          "Multiple Customers","Various",               2),

    ("uc_am_sw_spend",    "AB InBev Saves $261k on PowerBI Licenses",                          "https://nexthink.com/blog/5-times-nexthink-helped-customers-reduce-it-costs",                                          "AB InBev",          "Consumer Goods",        1),
    ("uc_am_sw_spend",    "Cleaning House: How One IT Team Saved $1.8M on SaaS Licensing",     "https://nexthink.com/blog/save-saas-licensing-app-experience/",                                                         "Anonymous",         "Various",               2),
    ("uc_am_sw_spend",    "Honeywell: Running a Cost-Efficient Workplace",                     "https://nexthink.com/resource/honeywell-cost-efficient-workplace",                                                      "Honeywell",         "Manufacturing",         3),

    ("uc_am_ai_optim",    "Secure AI Adoption at Scale with Nexthink Adopt",                   "https://nexthink.com/resource/empower-smarter-safer-ai-adoption-across-your-enterprise",                               "Anonymous",         "Utilities",             1),
    ("uc_am_ai_optim",    "How Leidos Is Building a Proactive, AI-Enabled IT Organization",    "https://nexthink.com/resource/how-leidos-is-building-a-proactive-ai-enabled-it-organization",                         "Leidos",            "Defense & Technology",  2),

    ("uc_am_vdi",         "The Virtual Experience: 3 Future-Facing VDI Use Cases",             "https://nexthink.com/en/blog/future-vdi-use-cases",                                                                     "Anonymous",         "Various",               1),

    # ── Sustainability ──
    ("uc_sus_carbon",     "How to Optimize Energy Consumption and Reduce Carbon Emissions",    "https://nexthink.com/blog/how-to-optimize-energy-consumption-and-carbon-emissions-with-nexthink",                      "Anonymous",         "Various",               1),

    ("uc_sus_energy",     "How to Optimize Energy Consumption and Reduce Carbon Emissions",    "https://nexthink.com/blog/how-to-optimize-energy-consumption-and-carbon-emissions-with-nexthink",                      "Anonymous",         "Various",               1),
    ("uc_sus_energy",     "5 Times Nexthink Helped Customers Reduce IT Costs",                 "https://nexthink.com/blog/5-times-nexthink-helped-customers-reduce-it-costs",                                          "Multiple Customers","Various",               2),

    ("uc_sus_hw_proc",    "Hospital Saves $900k on Hardware Refresh",                          "https://nexthink.com/blog/5-times-nexthink-helped-customers-reduce-it-costs",                                          "Anonymous",         "Healthcare",            1),
    ("uc_sus_hw_proc",    "Extend Device Lifecycles and Increase Employee Happiness",          "https://nexthink.com/blog/extend-hardware-refresh-cycle",                                                               "Anonymous",         "Various",               2),

    # ── Strategy ──
    ("uc_str_experience", "How Campbell's Transforms its Digital Workplace with DEX Score",    "https://nexthink.com/blog/how-to-measure-digital-employee-experience",                                                  "Campbell's",        "Consumer Goods",        1),
    ("uc_str_experience", "Johnson & Johnson: A Frictionless IT Experience",                   "https://nexthink.com/resource/johnson-johnson-a-frictionless-it-experience",                                           "Johnson & Johnson", "Healthcare",            2),
    ("uc_str_experience", "Fortune 500 Healthcare Cracks Their IT Problems Once and For All",  "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            3),

    ("uc_str_revenue",    "Honeywell: Running a Cost-Efficient Workplace",                     "https://nexthink.com/resource/honeywell-cost-efficient-workplace",                                                      "Honeywell",         "Manufacturing",         1),
    ("uc_str_revenue",    "A Year in the Life of a Nexthink Customer",                        "https://nexthink.com/en/resource/a-year-in-the-life-of-a-nexthink-customer",                                           "Anonymous",         "Financial Services",    2),

    ("uc_str_data",       "AXA Unlocks the Power of Enterprise IT Visibility",                 "https://nexthink.com/en/resource/axa-unlocks-the-power-of-enterprise-it-visibility",                                   "AXA",               "Insurance",             1),
    ("uc_str_data",       "A Year in the Life of a Nexthink Customer",                        "https://nexthink.com/en/resource/a-year-in-the-life-of-a-nexthink-customer",                                           "Anonymous",         "Financial Services",    2),

    ("uc_str_talent",     "Johnson & Johnson: A Frictionless IT Experience",                   "https://nexthink.com/resource/johnson-johnson-a-frictionless-it-experience",                                           "Johnson & Johnson", "Healthcare",            1),
    ("uc_str_talent",     "Fortune 500 Healthcare Cracks Their IT Problems Once and For All",  "https://nexthink.com/resource/fortune-500-healthcare-case-study",                                                      "Anonymous",         "Healthcare",            2),

    ("uc_str_xla",        "How Campbell's Transforms its Digital Workplace with DEX Score",    "https://nexthink.com/blog/how-to-measure-digital-employee-experience",                                                  "Campbell's",        "Consumer Goods",        1),
    ("uc_str_xla",        "A Year in the Life of a Nexthink Customer",                        "https://nexthink.com/en/resource/a-year-in-the-life-of-a-nexthink-customer",                                           "Anonymous",         "Financial Services",    2),

    # ── Change & Transformation ──
    ("uc_ct_budget",      "Honeywell: Running a Cost-Efficient Workplace",                     "https://nexthink.com/resource/honeywell-cost-efficient-workplace",                                                      "Honeywell",         "Manufacturing",         1),
    ("uc_ct_budget",      "German Manufacturer Saves $1.6M on Windows Migration",             "https://nexthink.com/blog/migration-savings/",                                                                          "Anonymous",         "Manufacturing",         2),

    ("uc_ct_transform",   "German Manufacturer Saves $1.6M on Windows Migration",             "https://nexthink.com/blog/migration-savings/",                                                                          "Anonymous",         "Manufacturing",         1),
    ("uc_ct_transform",   "Leidos Builds a Proactive IT Strategy with Nexthink",              "https://nexthink.com/resource/leidos-builds-a-proactive-it-strategy-with-nexthink",                                   "Leidos",            "Defense & Technology",  2),

    ("uc_ct_ai_adopt",    "Secure AI Adoption at Scale with Nexthink Adopt",                   "https://nexthink.com/resource/empower-smarter-safer-ai-adoption-across-your-enterprise",                               "Anonymous",         "Utilities",             1),
    ("uc_ct_ai_adopt",    "How Leidos Is Building a Proactive, AI-Enabled IT Organization",    "https://nexthink.com/resource/how-leidos-is-building-a-proactive-ai-enabled-it-organization",                         "Leidos",            "Defense & Technology",  2),

    ("uc_ct_mna",         "How Honeywell Achieved M&A Efficiency with DEX Solutions",          "https://nexthink.com/blog/how-honeywell-has-achieved-manda-efficiency-with-dex-solutions",                             "Honeywell",         "Manufacturing",         1),
    ("uc_ct_mna",         "German Manufacturer Saves $1.6M on Windows Migration",             "https://nexthink.com/blog/migration-savings/",                                                                          "Anonymous",         "Manufacturing",         2),
]

# ── Feature assignments ────────────────────────────────────────────────────────
UC_FEATURES = [
    ("uc_sm_reactive",    "f_collect_endpoint"),
    ("uc_sm_reactive",    "f_corr_ai"),
    ("uc_sm_reactive",    "f_act_dashboards"),
    ("uc_sm_reactive",    "f_alert"),
    ("uc_sm_reactive",    "f_diagnostic"),
    ("uc_sm_automation",  "f_collect_endpoint"),
    ("uc_sm_automation",  "f_corr_ai"),
    ("uc_sm_automation",  "f_act_remediation"),
    ("uc_sm_automation",  "f_act_campaign"),
    ("uc_sm_automation",  "f_library"),
    ("uc_sm_efficiency",  "f_collect_endpoint"),
    ("uc_sm_efficiency",  "f_corr_ai"),
    ("uc_sm_efficiency",  "f_act_dashboards"),
    ("uc_sm_efficiency",  "f_investigations"),
    ("uc_sm_efficiency",  "f_act_api"),
    ("uc_sm_prevention",  "f_collect_endpoint"),
    ("uc_sm_prevention",  "f_corr_ai"),
    ("uc_sm_prevention",  "f_act_dashboards"),
    ("uc_sm_prevention",  "f_alert"),
    ("uc_sm_prevention",  "f_act_remediation"),
    ("uc_sm_onboarding",  "f_collect_endpoint"),
    ("uc_sm_onboarding",  "f_act_campaign"),
    ("uc_sm_onboarding",  "f_collect_sentiment"),
    ("uc_sm_onboarding",  "f_library"),
    ("uc_spark",          "f_spark"),
    ("uc_spark",          "f_collect_endpoint"),
    ("uc_spark",          "f_corr_ai"),
    ("uc_spark",          "f_amplify"),
    ("uc_spark",          "f_act_remediation"),
    ("uc_spark",          "f_workspace"),

    ("uc_comp_sec_posture","f_collect_endpoint"),
    ("uc_comp_sec_posture","f_corr_ai"),
    ("uc_comp_sec_posture","f_act_dashboards"),
    ("uc_comp_sec_posture","f_act_remediation"),
    ("uc_comp_sec_posture","f_act_api"),
    ("uc_comp_standards",  "f_collect_endpoint"),
    ("uc_comp_standards",  "f_act_remediation"),
    ("uc_comp_standards",  "f_act_dashboards"),
    ("uc_comp_standards",  "f_gov_rbac"),
    ("uc_comp_human_risk", "f_collect_endpoint"),
    ("uc_comp_human_risk", "f_corr_ai"),
    ("uc_comp_human_risk", "f_act_dashboards"),
    ("uc_comp_human_risk", "f_collect_sentiment"),
    ("uc_comp_data",       "f_collect_endpoint"),
    ("uc_comp_data",       "f_act_api"),
    ("uc_comp_data",       "f_gov_rbac"),
    ("uc_comp_data",       "f_act_dashboards"),
    ("uc_comp_ai",         "f_ai_tools"),
    ("uc_comp_ai",         "f_collect_endpoint"),
    ("uc_comp_ai",         "f_act_dashboards"),
    ("uc_comp_ai",         "f_gov_rbac"),

    ("uc_am_hw_refresh",   "f_collect_endpoint"),
    ("uc_am_hw_refresh",   "f_corr_ai"),
    ("uc_am_hw_refresh",   "f_act_dashboards"),
    ("uc_am_hw_refresh",   "f_exp_central"),
    ("uc_am_hw_sizing",    "f_collect_endpoint"),
    ("uc_am_hw_sizing",    "f_corr_ai"),
    ("uc_am_hw_sizing",    "f_act_dashboards"),
    ("uc_am_hw_sizing",    "f_act_api"),
    ("uc_am_sw_spend",     "f_collect_endpoint"),
    ("uc_am_sw_spend",     "f_collect_external"),
    ("uc_am_sw_spend",     "f_act_dashboards"),
    ("uc_am_sw_spend",     "f_act_api"),
    ("uc_am_ai_optim",     "f_ai_tools"),
    ("uc_am_ai_optim",     "f_collect_endpoint"),
    ("uc_am_ai_optim",     "f_act_dashboards"),
    ("uc_am_ai_optim",     "f_act_api"),
    ("uc_am_vdi",          "f_vdi"),
    ("uc_am_vdi",          "f_collect_endpoint"),
    ("uc_am_vdi",          "f_corr_ai"),
    ("uc_am_vdi",          "f_act_dashboards"),
    ("uc_am_mobile",       "f_mobile"),
    ("uc_am_mobile",       "f_collect_endpoint"),
    ("uc_am_mobile",       "f_act_dashboards"),

    ("uc_sus_carbon",      "f_collect_endpoint"),
    ("uc_sus_carbon",      "f_act_dashboards"),
    ("uc_sus_carbon",      "f_act_api"),
    ("uc_sus_carbon",      "f_act_campaign"),
    ("uc_sus_energy",      "f_collect_endpoint"),
    ("uc_sus_energy",      "f_act_remediation"),
    ("uc_sus_energy",      "f_act_dashboards"),
    ("uc_sus_hw_proc",     "f_collect_endpoint"),
    ("uc_sus_hw_proc",     "f_corr_ai"),
    ("uc_sus_hw_proc",     "f_act_dashboards"),
    ("uc_sus_hw_proc",     "f_act_api"),

    ("uc_str_experience",  "f_collect_endpoint"),
    ("uc_str_experience",  "f_collect_sentiment"),
    ("uc_str_experience",  "f_corr_dex_score"),
    ("uc_str_experience",  "f_exp_central"),
    ("uc_str_experience",  "f_act_dashboards"),
    ("uc_str_revenue",     "f_collect_endpoint"),
    ("uc_str_revenue",     "f_corr_ai"),
    ("uc_str_revenue",     "f_act_dashboards"),
    ("uc_str_revenue",     "f_act_api"),
    ("uc_str_data",        "f_collect_endpoint"),
    ("uc_str_data",        "f_act_dashboards"),
    ("uc_str_data",        "f_act_api"),
    ("uc_str_data",        "f_investigations"),
    ("uc_str_talent",      "f_collect_sentiment"),
    ("uc_str_talent",      "f_corr_dex_score"),
    ("uc_str_talent",      "f_exp_central"),
    ("uc_str_talent",      "f_act_campaign"),
    ("uc_str_xla",         "f_collect_endpoint"),
    ("uc_str_xla",         "f_collect_sentiment"),
    ("uc_str_xla",         "f_corr_dex_score"),
    ("uc_str_xla",         "f_exp_central"),
    ("uc_str_xla",         "f_act_dashboards"),

    ("uc_ct_budget",       "f_collect_endpoint"),
    ("uc_ct_budget",       "f_act_dashboards"),
    ("uc_ct_budget",       "f_act_api"),
    ("uc_ct_budget",       "f_investigations"),
    ("uc_ct_transform",    "f_collect_endpoint"),
    ("uc_ct_transform",    "f_corr_ai"),
    ("uc_ct_transform",    "f_act_campaign"),
    ("uc_ct_transform",    "f_act_dashboards"),
    ("uc_ct_transform",    "f_library"),
    ("uc_ct_ai_adopt",     "f_ai_tools"),
    ("uc_ct_ai_adopt",     "f_collect_endpoint"),
    ("uc_ct_ai_adopt",     "f_act_campaign"),
    ("uc_ct_ai_adopt",     "f_act_dashboards"),
    ("uc_ct_mna",          "f_collect_endpoint"),
    ("uc_ct_mna",          "f_collect_external"),
    ("uc_ct_mna",          "f_corr_ai"),
    ("uc_ct_mna",          "f_act_dashboards"),
    ("uc_ct_mna",          "f_act_api"),
]

# ── Persona assignments ────────────────────────────────────────────────────────
UC_PERSONAS = [
    ("uc_sm_reactive",    "p_servicedesk"),
    ("uc_sm_reactive",    "p_workplace"),
    ("uc_sm_automation",  "p_servicedesk"),
    ("uc_sm_automation",  "p_workplace"),
    ("uc_sm_efficiency",  "p_servicedesk"),
    ("uc_sm_efficiency",  "p_workplace"),
    ("uc_sm_efficiency",  "p_cio"),
    ("uc_sm_prevention",  "p_workplace"),
    ("uc_sm_prevention",  "p_servicedesk"),
    ("uc_sm_onboarding",  "p_servicedesk"),
    ("uc_sm_onboarding",  "p_hr"),
    ("uc_spark",          "p_servicedesk"),
    ("uc_spark",          "p_workplace"),
    ("uc_spark",          "p_finance"),

    ("uc_comp_sec_posture","p_security"),
    ("uc_comp_sec_posture","p_workplace"),
    ("uc_comp_standards",  "p_security"),
    ("uc_comp_standards",  "p_workplace"),
    ("uc_comp_human_risk", "p_security"),
    ("uc_comp_human_risk", "p_hr"),
    ("uc_comp_data",       "p_security"),
    ("uc_comp_data",       "p_workplace"),
    ("uc_comp_data",       "p_cio"),
    ("uc_comp_ai",         "p_security"),
    ("uc_comp_ai",         "p_cio"),

    ("uc_am_hw_refresh",   "p_finance"),
    ("uc_am_hw_refresh",   "p_workplace"),
    ("uc_am_hw_sizing",    "p_finance"),
    ("uc_am_hw_sizing",    "p_workplace"),
    ("uc_am_sw_spend",     "p_finance"),
    ("uc_am_sw_spend",     "p_workplace"),
    ("uc_am_ai_optim",     "p_finance"),
    ("uc_am_ai_optim",     "p_cio"),
    ("uc_am_vdi",          "p_workplace"),
    ("uc_am_vdi",          "p_finance"),
    ("uc_am_mobile",       "p_workplace"),
    ("uc_am_mobile",       "p_finance"),

    ("uc_sus_carbon",      "p_cio"),
    ("uc_sus_carbon",      "p_finance"),
    ("uc_sus_energy",      "p_finance"),
    ("uc_sus_energy",      "p_workplace"),
    ("uc_sus_hw_proc",     "p_finance"),
    ("uc_sus_hw_proc",     "p_workplace"),

    ("uc_str_experience",  "p_cio"),
    ("uc_str_experience",  "p_hr"),
    ("uc_str_revenue",     "p_cio"),
    ("uc_str_revenue",     "p_finance"),
    ("uc_str_data",        "p_cio"),
    ("uc_str_data",        "p_workplace"),
    ("uc_str_talent",      "p_hr"),
    ("uc_str_talent",      "p_cio"),
    ("uc_str_xla",         "p_cio"),
    ("uc_str_xla",         "p_hr"),
    ("uc_str_xla",         "p_workplace"),

    ("uc_ct_budget",       "p_cio"),
    ("uc_ct_budget",       "p_finance"),
    ("uc_ct_transform",    "p_cio"),
    ("uc_ct_transform",    "p_workplace"),
    ("uc_ct_ai_adopt",     "p_cio"),
    ("uc_ct_ai_adopt",     "p_workplace"),
    ("uc_ct_mna",          "p_cio"),
    ("uc_ct_mna",          "p_finance"),
]

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    if not WB_PATH.exists():
        raise SystemExit(f"Not found: {WB_PATH}")
    wb = load_workbook(WB_PATH)

    # value_categories
    ws = wb["value_categories"]
    write_sheet(ws,
        ["id", "name", "color", "tagline", "order"],
        [(r[0], r[1], r[2], r[3], r[4]) for r in VALUE_CATEGORIES],
        "The 6 buckets of value. Each use case belongs to exactly one of these.",
    )

    # use_cases — id, name, vc_id, short_pitch, customer fields, value_description, order
    ws = wb["use_cases"]
    write_sheet(ws,
        ["id", "name", "value_category_id", "short_pitch",
         "customer_name", "customer_industry", "customer_oneliner", "customer_result",
         "value_description", "order"],
        [(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9]) for r in USE_CASES],
        "The atomic unit. Belongs to exactly one value_category_id.",
    )

    # use_case_stories (new or existing sheet)
    if "use_case_stories" not in wb.sheetnames:
        wb.create_sheet("use_case_stories")
    ws = wb["use_case_stories"]
    write_sheet(ws,
        ["use_case_id", "title", "url", "customer", "industry", "order"],
        list(UC_STORIES),
        "Up to 3 public customer stories per use case. All URLs are from nexthink.com.",
    )

    # use_case_benefits (new or existing sheet)
    if "use_case_benefits" not in wb.sheetnames:
        wb.create_sheet("use_case_benefits")
    ws = wb["use_case_benefits"]
    write_sheet(ws,
        ["use_case_id", "benefit_type"],
        list(UC_BENEFITS),
        f"M:N: IT & Business benefits each use case delivers. benefit_type must be one of: {', '.join(ALL_BENEFITS)}",
    )

    # use_case_features
    ws = wb["use_case_features"]
    write_sheet(ws,
        ["use_case_id", "feature_id"],
        list(UC_FEATURES),
        "Many-to-many: which foundation_features each use case relies on. One row per pair.",
    )

    # use_case_personas
    ws = wb["use_case_personas"]
    write_sheet(ws,
        ["use_case_id", "persona_id"],
        list(UC_PERSONAS),
        "Many-to-many: which personas each use case targets.",
    )

    # maturity_levels — clear content but keep sheet
    ws = wb["maturity_levels"]
    ws.delete_rows(1, ws.max_row + 1)
    ws.cell(1, 1, "Maturity levels are no longer used. This sheet is kept for reference only.").font = note_style()

    # use_case_metrics — clear content but keep sheet (replaced by value_description)
    if "use_case_metrics" in wb.sheetnames:
        ws = wb["use_case_metrics"]
        ws.delete_rows(1, ws.max_row + 1)
        ws.cell(1, 1, "use_case_metrics replaced by value_description column in use_cases sheet.").font = note_style()

    wb.save(WB_PATH)
    print(f"OK → {WB_PATH} written")
    print(f"  {len(VALUE_CATEGORIES)} value categories")
    print(f"  {len(USE_CASES)} use cases (with value_description)")
    print(f"  {len(UC_STORIES)} story links")
    print(f"  {len(UC_BENEFITS)} benefit links")
    print(f"  {len(UC_FEATURES)} feature links")
    print(f"  {len(UC_PERSONAS)} persona links")

if __name__ == "__main__":
    main()

"""Builds the master content workbook for the SaaS platform infographic.
Each sheet maps to one entity; relations live in dedicated join sheets."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUT = "platform_content.xlsx"

HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
NOTE_FILL   = PatternFill("solid", start_color="FEF3C7")
ZEBRA_FILL  = PatternFill("solid", start_color="F8FAFC")
BODY_FONT   = Font(name="Arial", size=10)
THIN        = Side(style="thin", color="E5E7EB")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def make_sheet(wb, name, headers, rows, widths=None, notes=None, table_style="TableStyleMedium2"):
    ws = wb.create_sheet(name)
    # Optional notes rows above the header
    note_rows = 0
    if notes:
        for n in notes:
            ws.cell(row=note_rows+1, column=1, value=n).font = Font(name="Arial", italic=True, size=9, color="6B7280")
            ws.cell(row=note_rows+1, column=1).fill = NOTE_FILL
            ws.merge_cells(start_row=note_rows+1, start_column=1, end_row=note_rows+1, end_column=max(1,len(headers)))
            note_rows += 1
    header_row = note_rows + 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 22
    for ri, row in enumerate(rows, header_row+1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if (ri - header_row) % 2 == 0:
                cell.fill = ZEBRA_FILL
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row+1, column=1)
    return ws

def main():
    wb = Workbook()
    wb.remove(wb.active)

    # ---------- README ----------
    ws = wb.create_sheet("README")
    lines = [
        ("Platform content workbook", True, 16, "1F2937"),
        ("Single source of truth for the interactive infographic.", False, 11, "374151"),
        ("", False, 10, "374151"),
        ("How to use", True, 13, "1F2937"),
        ("1. Edit the sheets below — never the generated data.json directly.", False, 10, "374151"),
        ("2. Each row needs a stable, lowercase, hyphen-free ID (e.g. uc_vdi_login_speed).", False, 10, "374151"),
        ("3. After editing, run:  python xlsx_to_json.py platform_content.xlsx data.json", False, 10, "374151"),
        ("4. Refresh the HTML — it picks up the new data.json automatically.", False, 10, "374151"),
        ("", False, 10, "374151"),
        ("Sheets", True, 13, "1F2937"),
        ("data_sources         — every system that feeds the platform", False, 10, "374151"),
        ("foundation_features  — capabilities of the platform itself (collect / correlate / act / govern)", False, 10, "374151"),
        ("value_categories     — the 6 logical groupings of business value", False, 10, "374151"),
        ("use_cases            — individual use cases (1 belongs to 1 value_category)", False, 10, "374151"),
        ("use_case_features    — join: which features each use case relies on (M:N)", False, 10, "374151"),
        ("personas             — user/buyer profiles", False, 10, "374151"),
        ("use_case_personas    — join: which personas care about each use case (M:N)", False, 10, "374151"),
        ("maturity_levels      — 1-row-per (use_case, level) describing the climb from L1→L4", False, 10, "374151"),
        ("", False, 10, "374151"),
        ("Conventions", True, 13, "1F2937"),
        ("• IDs are immutable — change the name freely, never the id (it would break links).", False, 10, "374151"),
        ("• Use the 'order' column to control display order; lower numbers come first.", False, 10, "374151"),
        ("• Multi-line text is fine; the renderer respects line breaks.", False, 10, "374151"),
    ]
    for i, (text, bold, size, color) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", bold=bold, size=size, color=color)
    ws.column_dimensions["A"].width = 110

    # ---------- data_sources ----------
    make_sheet(wb, "data_sources",
        headers=["id", "name", "category", "icon", "description", "order"],
        notes=["Every device or external system that feeds telemetry into the platform.",
               "category is free-form (e.g. 'endpoint', 'virtual', 'mobile', 'external'). icon is a short keyword used by the UI."],
        rows=[
            ["src_vdi",     "Virtual sessions",   "virtual",  "vdi",   "Citrix, VMware, Azure Virtual Desktop, Cloud PCs.", 10],
            ["src_laptop",  "Laptops",            "endpoint", "laptop","Windows, macOS, Linux laptops — full hardware/software/network telemetry.", 20],
            ["src_desktop", "Desktops",           "endpoint", "desktop","Office and shop-floor workstations and kiosks.", 30],
            ["src_mobile",  "Mobile devices",     "mobile",   "mobile","iOS, Android phones and tablets — connectivity & app health.", 40],
            ["src_external","External systems",   "external", "plug",  "ITSM, IdP, MDM, network, SaaS apps — connectors enrich device truth.", 50],
            ["src_sentiment","User sentiment",    "external", "smile", "In-context surveys correlating how users feel with what we measure.", 60],
        ],
        widths=[24, 22, 14, 12, 70, 10])

    # ---------- foundation_features ----------
    make_sheet(wb, "foundation_features",
        headers=["id", "name", "group", "description", "order"],
        notes=["Capabilities offered by the platform. group must be one of: collect, correlate, act, govern.",
               "These are referenced by use_cases via the use_case_features sheet."],
        rows=[
            ["f_collect_endpoint",  "Endpoint telemetry",        "collect",   "Lightweight agent collecting hardware, OS, app, network and crash signals.", 10],
            ["f_collect_virtual",   "Virtual session telemetry", "collect",   "Login times, latency, app responsiveness inside Citrix/VMware/AVD.", 20],
            ["f_collect_sentiment", "Employee sentiment surveys","collect",   "In-context micro-surveys triggered by events or schedule.", 30],
            ["f_collect_external",  "External connectors",       "collect",   "Pull telemetry & context from ITSM, IdP, MDM, network and SaaS.", 40],
            ["f_corr_realtime",     "Realtime data fabric",      "correlate", "Stream processing engine that joins device, user, app and network data.", 50],
            ["f_corr_ai",           "AI insights & anomaly",     "correlate", "ML detecting issues, drift, performance regressions across the fleet.", 60],
            ["f_corr_dex_score",    "DEX scoring engine",        "correlate", "Composite digital experience score per user / device / location.", 70],
            ["f_act_dashboards",    "Dashboards & analytics",    "act",       "Pre-built and custom dashboards for every persona.", 80],
            ["f_act_remediation",   "Automated remediation",     "act",       "Triggered or scheduled scripts/workflows fixing issues at scale.", 90],
            ["f_act_campaign",      "Employee campaigns",        "act",       "Targeted in-app guidance, nudges and self-help to end users.", 100],
            ["f_act_api",           "Open APIs & webhooks",      "act",       "Push insights into ITSM, SIEM, BI and any downstream system.", 110],
            ["f_gov_compliance",    "Policy & compliance engine","govern",    "Continuous compliance checks against configurable baselines.", 120],
            ["f_gov_rbac",          "RBAC & data residency",     "govern",    "Role-based access, regional data hosting, audit trail.", 130],
        ],
        widths=[24, 30, 12, 70, 10])

    # ---------- value_categories ----------
    make_sheet(wb, "value_categories",
        headers=["id", "name", "color", "tagline", "order"],
        notes=["The 6 buckets of value. Each use case belongs to exactly one of these.",
               "color is a hex code used in the visualization (without the #)."],
        rows=[
            ["vc_dex",       "Digital employee experience", "22D3EE", "Measure & improve how work feels", 10],
            ["vc_proactive", "Proactive IT operations",     "6366F1", "Detect and fix before users notice", 20],
            ["vc_change",    "Change & transformation",     "A78BFA", "De-risk every migration & rollout", 30],
            ["vc_security",  "Security & compliance",       "F472B6", "Continuous posture across endpoints", 40],
            ["vc_sustain",   "Sustainability & FinOps",     "34D399", "Cut waste, energy and IT spend",   50],
            ["vc_workplace", "Workplace & productivity",    "FBBF24", "Modern workplace adoption at scale",60],
        ],
        widths=[18, 30, 10, 50, 10])

    # ---------- use_cases ----------
    make_sheet(wb, "use_cases",
        headers=["id", "name", "value_category_id", "short_pitch",
                 "value_metric", "value_formula", "value_example",
                 "customer_name", "customer_industry", "customer_oneliner", "customer_result", "order"],
        notes=["The atomic unit. Belongs to exactly one value_category_id.",
               "value_formula is plain text describing how the ROI is computed; value_example is a worked number."],
        rows=[
            ["uc_dex_score", "Company-wide DEX score",          "vc_dex",
             "One number that reflects how every employee experiences IT.",
             "DEX score uplift",
             "(target_score - baseline_score) × population × hours_saved_per_point",
             "+12 pts × 10,000 users × 1.5h/yr ≈ 180,000 h/yr regained.",
             "Acme Corp", "Manufacturing",
             "Rolled out DEX scoring across 18 countries in 6 weeks.",
             "+12 DEX points in 5 months · 30% fewer P1 tickets.", 10],
            ["uc_dex_persona", "Persona-based experience analytics", "vc_dex",
             "Compare experience by role, location, hardware, app — and act on outliers.",
             "Hours of friction avoided",
             "users_in_persona × friction_minutes_saved / 60",
             "2,300 sales reps × 22 min/day saved ≈ 14,000 h/yr.",
             "Globex Bank", "Financial services",
             "Identified that traders on a specific laptop model were 4x slower.",
             "Replaced 1,200 devices proactively, saved $2.1M in support cost.", 20],
            ["uc_proactive_fleet", "Fleet-wide issue detection",  "vc_proactive",
             "Spot performance regressions and crash patterns across the whole estate.",
             "Tickets avoided",
             "incidents_detected × avg_users_impacted × deflection_rate",
             "60 issues × 800 users × 70% ≈ 33,600 tickets avoided/yr.",
             "Initech", "Retail",
             "Caught a faulty driver rollout 3 hours after deployment.",
             "Rolled back to 9,000 endpoints in 25 minutes — zero tickets opened.", 10],
            ["uc_proactive_l1_deflect", "L1 ticket deflection",  "vc_proactive",
             "Self-heal common issues automatically before users open a ticket.",
             "Cost of a ticket",
             "tickets_deflected × cost_per_ticket",
             "120,000 deflections × $18 ≈ $2.16M/yr.",
             "Umbrella", "Healthcare",
             "Replaced 14 standard runbooks with auto-remediation.",
             "L1 ticket volume −38% in 4 months.", 20],
            ["uc_change_win11", "Windows 11 migration",          "vc_change",
             "Continuous readiness, pilot tracking and post-deploy validation.",
             "Migration project cost",
             "users × project_h_saved × loaded_rate",
             "20k users × 1.2h saved × $80 ≈ $1.92M.",
             "Stark Industries", "Aerospace",
             "Migrated 65,000 endpoints to Windows 11 in 4 months.",
             "98% on-time migration · zero P1 escalations post-deploy.", 10],
            ["uc_change_m365", "M365 adoption & rollout",        "vc_change",
             "Track adoption of Teams/Outlook/OneDrive and nudge laggards.",
             "License waste avoided",
             "unused_licenses × license_cost",
             "4,500 unused × $22 ≈ $99k/mo reclaimed.",
             "Wayne Enterprises", "Conglomerate",
             "Identified 4,500 dormant E5 licenses in week 2.",
             "Reclaimed $1.2M annual SaaS spend.", 20],
            ["uc_sec_posture", "Endpoint security posture",      "vc_security",
             "Continuous compliance with configurable baselines, with drift alerts.",
             "% endpoints compliant",
             "compliant_endpoints / total_endpoints",
             "From 78% → 96% compliance in 90 days.",
             "Cyberdyne", "Defense",
             "Used policy engine to harden 22,000 endpoints to ISO 27001.",
             "Audit prep time: 6 weeks → 4 days.", 10],
            ["uc_sec_vuln_response", "Rapid vulnerability response", "vc_security",
             "From CVE disclosure to patched fleet in hours, not weeks.",
             "Mean time to remediate",
             "current_MTTR - new_MTTR",
             "MTTR 14 days → 36 hours on critical CVEs.",
             "Tyrell Corp", "Tech",
             "Patched Log4j across 30k endpoints in 11 hours.",
             "Avoided 2 audit findings · zero downtime.", 20],
            ["uc_sustain_energy", "Energy & CO₂ reduction",      "vc_sustain",
             "Right-size sleep policies, identify always-on devices, report on CO₂.",
             "kWh / year saved",
             "devices × hours_optimized × wattage",
             "30k × 4h × 35W ≈ 1.5 GWh/yr.",
             "Greenfield Co", "Energy",
             "Cut idle laptop power by 38% via policy + nudges.",
             "−540 t CO₂eq/yr · ESG report data automated.", 10],
            ["uc_sustain_lifecycle", "Hardware lifecycle extension","vc_sustain",
             "Identify devices that can be kept longer based on real performance.",
             "CapEx avoided",
             "devices_kept × replacement_cost",
             "5,000 devices × $1,200 ≈ $6M/yr CapEx avoided.",
             "EverGreen Bank", "Financial services",
             "Extended laptop lifecycle from 3 to 4.5 years.",
             "$8.4M CapEx deferred over 2 years.", 20],
            ["uc_workplace_collab", "Collaboration tools health", "vc_workplace",
             "Detect call quality, meeting drops, and Teams/Zoom issues by site.",
             "Call quality score",
             "good_calls / total_calls",
             "82% → 96% good calls in 3 months.",
             "MegaCorp", "Pharma",
             "Mapped poor call quality to 7 sites with bad Wi-Fi.",
             "Fixed root cause — exec NPS +18.", 10],
            ["uc_workplace_apps", "App rationalization",         "vc_workplace",
             "See which apps are actually used and consolidate the rest.",
             "SaaS spend reclaimed",
             "unused_seats × seat_cost",
             "12k unused seats × $12 ≈ $144k/mo.",
             "Pied Piper", "Tech",
             "Cut active app catalog from 420 to 230.",
             "$3.8M annual SaaS spend reclaimed.", 20],
        ],
        widths=[26, 32, 18, 50, 22, 42, 42, 18, 16, 42, 42, 8])

    # ---------- use_case_features (M:N) ----------
    make_sheet(wb, "use_case_features",
        headers=["use_case_id", "feature_id"],
        notes=["Many-to-many: which foundation_features each use case relies on. One row per pair."],
        rows=[
            # DEX
            ["uc_dex_score",        "f_collect_endpoint"],
            ["uc_dex_score",        "f_collect_virtual"],
            ["uc_dex_score",        "f_collect_sentiment"],
            ["uc_dex_score",        "f_corr_dex_score"],
            ["uc_dex_score",        "f_act_dashboards"],
            ["uc_dex_persona",      "f_collect_endpoint"],
            ["uc_dex_persona",      "f_corr_realtime"],
            ["uc_dex_persona",      "f_corr_dex_score"],
            ["uc_dex_persona",      "f_act_dashboards"],
            # Proactive
            ["uc_proactive_fleet",  "f_collect_endpoint"],
            ["uc_proactive_fleet",  "f_corr_realtime"],
            ["uc_proactive_fleet",  "f_corr_ai"],
            ["uc_proactive_fleet",  "f_act_dashboards"],
            ["uc_proactive_l1_deflect","f_collect_endpoint"],
            ["uc_proactive_l1_deflect","f_corr_ai"],
            ["uc_proactive_l1_deflect","f_act_remediation"],
            ["uc_proactive_l1_deflect","f_act_campaign"],
            # Change
            ["uc_change_win11",     "f_collect_endpoint"],
            ["uc_change_win11",     "f_collect_external"],
            ["uc_change_win11",     "f_corr_realtime"],
            ["uc_change_win11",     "f_act_dashboards"],
            ["uc_change_win11",     "f_act_campaign"],
            ["uc_change_m365",      "f_collect_endpoint"],
            ["uc_change_m365",      "f_collect_external"],
            ["uc_change_m365",      "f_act_campaign"],
            ["uc_change_m365",      "f_act_api"],
            # Security
            ["uc_sec_posture",      "f_collect_endpoint"],
            ["uc_sec_posture",      "f_gov_compliance"],
            ["uc_sec_posture",      "f_act_remediation"],
            ["uc_sec_posture",      "f_act_dashboards"],
            ["uc_sec_vuln_response","f_collect_endpoint"],
            ["uc_sec_vuln_response","f_corr_ai"],
            ["uc_sec_vuln_response","f_act_remediation"],
            ["uc_sec_vuln_response","f_act_api"],
            # Sustain
            ["uc_sustain_energy",   "f_collect_endpoint"],
            ["uc_sustain_energy",   "f_act_remediation"],
            ["uc_sustain_energy",   "f_act_campaign"],
            ["uc_sustain_energy",   "f_act_dashboards"],
            ["uc_sustain_lifecycle","f_collect_endpoint"],
            ["uc_sustain_lifecycle","f_corr_ai"],
            ["uc_sustain_lifecycle","f_act_dashboards"],
            # Workplace
            ["uc_workplace_collab", "f_collect_endpoint"],
            ["uc_workplace_collab", "f_collect_external"],
            ["uc_workplace_collab", "f_corr_realtime"],
            ["uc_workplace_collab", "f_act_dashboards"],
            ["uc_workplace_apps",   "f_collect_endpoint"],
            ["uc_workplace_apps",   "f_collect_external"],
            ["uc_workplace_apps",   "f_act_dashboards"],
            ["uc_workplace_apps",   "f_act_api"],
        ],
        widths=[26, 28])

    # ---------- personas ----------
    make_sheet(wb, "personas",
        headers=["id", "name", "role", "icon", "order"],
        notes=["The roles that consume insights from the platform."],
        rows=[
            ["p_cio",        "CIO / IT leadership",    "Strategy & outcomes",   "leader",  10],
            ["p_workplace",  "Workplace engineering",  "Build & run",           "engineer",20],
            ["p_servicedesk","Service desk",           "Resolve faster",        "support", 30],
            ["p_security",   "Security & risk",        "Posture & response",    "shield",  40],
            ["p_hr",         "HR & People",            "Employee experience",   "people",  50],
            ["p_finance",    "Finance & procurement",  "Spend & assets",        "coin",    60],
        ],
        widths=[18, 24, 26, 12, 8])

    # ---------- use_case_personas (M:N) ----------
    make_sheet(wb, "use_case_personas",
        headers=["use_case_id", "persona_id"],
        notes=["Many-to-many: which personas each use case targets."],
        rows=[
            ["uc_dex_score",        "p_cio"],
            ["uc_dex_score",        "p_workplace"],
            ["uc_dex_score",        "p_hr"],
            ["uc_dex_persona",      "p_workplace"],
            ["uc_dex_persona",      "p_servicedesk"],
            ["uc_proactive_fleet",  "p_workplace"],
            ["uc_proactive_fleet",  "p_servicedesk"],
            ["uc_proactive_l1_deflect","p_servicedesk"],
            ["uc_proactive_l1_deflect","p_finance"],
            ["uc_change_win11",     "p_cio"],
            ["uc_change_win11",     "p_workplace"],
            ["uc_change_m365",      "p_workplace"],
            ["uc_change_m365",      "p_finance"],
            ["uc_sec_posture",      "p_security"],
            ["uc_sec_posture",      "p_workplace"],
            ["uc_sec_vuln_response","p_security"],
            ["uc_sec_vuln_response","p_workplace"],
            ["uc_sustain_energy",   "p_cio"],
            ["uc_sustain_energy",   "p_finance"],
            ["uc_sustain_lifecycle","p_finance"],
            ["uc_workplace_collab", "p_workplace"],
            ["uc_workplace_collab", "p_servicedesk"],
            ["uc_workplace_apps",   "p_finance"],
            ["uc_workplace_apps",   "p_workplace"],
        ],
        widths=[26, 18])

    # ---------- maturity_levels ----------
    make_sheet(wb, "maturity_levels",
        headers=["use_case_id", "level", "label", "example"],
        notes=["Define 3-4 levels per use case. level is an integer 1..4."],
        rows=[
            ["uc_dex_score", 1, "Reactive",   "DEX measured ad-hoc through surveys only."],
            ["uc_dex_score", 2, "Visible",    "Continuous DEX score in a single dashboard."],
            ["uc_dex_score", 3, "Comparable", "Score broken down by persona, site, hardware."],
            ["uc_dex_score", 4, "Actioned",   "DEX is an exec KPI; actions auto-routed to owners."],

            ["uc_proactive_fleet", 1, "Manual",     "Issues found through tickets and pings."],
            ["uc_proactive_fleet", 2, "Detected",   "Anomalies surfaced in a dashboard."],
            ["uc_proactive_fleet", 3, "Investigated","Root-cause workflow with full context."],
            ["uc_proactive_fleet", 4, "Self-healing","Detection triggers automated remediation at scale."],

            ["uc_change_win11", 1, "Best effort",  "Migration tracked via spreadsheets."],
            ["uc_change_win11", 2, "Tracked",      "Readiness and progress dashboards in place."],
            ["uc_change_win11", 3, "Validated",    "Post-deploy validation per ring + auto-rollback."],
            ["uc_change_win11", 4, "Continuous",   "Every future change reuses the same pipeline."],

            ["uc_sec_posture", 1, "Snapshot",    "Quarterly compliance audits."],
            ["uc_sec_posture", 2, "Continuous",  "Always-on policy checks across the fleet."],
            ["uc_sec_posture", 3, "Drift-aware", "Drift triggers automated remediation."],
            ["uc_sec_posture", 4, "Audit-ready", "Live evidence pack for any auditor anytime."],

            ["uc_sustain_energy", 1, "Unknown",   "No measurement of endpoint energy."],
            ["uc_sustain_energy", 2, "Measured",  "kWh and CO₂ reported by site/device."],
            ["uc_sustain_energy", 3, "Optimized", "Sleep policies and nudges deployed."],
            ["uc_sustain_energy", 4, "Reported",  "Data feeds the corporate ESG report automatically."],

            ["uc_workplace_collab", 1, "Anecdotal", "Quality issues surfaced via complaints."],
            ["uc_workplace_collab", 2, "Quantified","Per-call quality metrics dashboard."],
            ["uc_workplace_collab", 3, "Localized", "Issues mapped to sites/Wi-Fi/devices."],
            ["uc_workplace_collab", 4, "Eliminated","Auto-routing to network/AV teams with SLAs."],
        ],
        widths=[26, 8, 16, 70])

    wb.save(OUT)
    print(f"wrote {OUT}")

if __name__ == "__main__":
    main()
